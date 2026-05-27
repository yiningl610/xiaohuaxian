import openpyxl

def load_data():
    # ── 读取花种换算 ──────────────────────────────────────
    seed_rates = {}
    flower_sources = {}
    flower_times = {}
    flower_prices = {}
    flower_units = {}
    wb = openpyxl.load_workbook("flowers.xlsx", data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        flower = str(data.get("花朵类型") or "").strip()
        rate = data.get("基础产量")
        source = str(data.get("获取方式") or "").strip()
        time_val = data.get("开花时间")
        price = data.get("价格")
        unit = str(data.get("单位") or "").strip()
        if flower and rate:
            seed_rates[flower] = float(rate)
            flower_sources[flower] = source if source else ""
            flower_times[flower] = int(time_val) if time_val else 0
            flower_prices[flower] = float(price) if price else 0
            flower_units[flower] = unit

    # ── 读取服装列表 ──────────────────────────────────────
    outfits_dict = {}
    wb2 = openpyxl.load_workbook("clothes.xlsx")
    ws2 = wb2.active
    headers2 = [cell.value.strip() for cell in ws2[1]]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers2, row))
        name = (data.get("套装") or "").strip()
        if not name:
            continue
        if name not in outfits_dict:
            outfits_dict[name] = {
                "name": name,
                "location": (data.get("兑换位置") or "").strip(),
                "parts": []
            }
        outfits_dict[name]["parts"].append({
            "part": (data.get("部件") or "").strip(),
            "flower": (data.get("所需花朵") or "").strip(),
            "count": int(data.get("所需数量") or 0)
        })

    # ── 按兑换位置分组 ────────────────────────────────────
    locations_dict = {}
    for outfit in outfits_dict.values():
        loc = outfit["location"]
        if loc not in locations_dict:
            locations_dict[loc] = []
        locations_dict[loc].append(outfit)

    locations = [{"location": loc, "outfits": o} for loc, o in locations_dict.items()]
    outfits = list(outfits_dict.values())

    return seed_rates, flower_sources, flower_times, flower_prices, flower_units, outfits, locations

def load_bag():
    bag_flowers = {}
    bag_seeds = {}
    wb = openpyxl.load_workbook("bag.xlsx")
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        flower = str(data.get("花朵类型") or "").strip()
        if flower:
            bag_flowers[flower] = int(data.get("花朵数量") or 0)
            bag_seeds[flower] = int(data.get("花种数量") or 0)
    return bag_flowers, bag_seeds